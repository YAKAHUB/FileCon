import io
import fitz  # PyMuPDF
from PIL import Image
import rarfile
import py7zr
from pydub import AudioSegment  # Requires ffmpeg
from docx import Document
import openpyxl
import moviepy.editor as mp
import magic  # For file type detection

# Supported categories and conversion options (no 'vector')
CONVERSION_OPTIONS = {
    'archive': {
        'zip': ['rar', '7z'],
        'rar': ['zip', '7z'],
        '7z': ['zip', 'rar']
    },
    'audio': {
        'mp3': ['wav', 'ogg'],
        'wav': ['mp3', 'ogg'],
        'ogg': ['mp3', 'wav']
    },
    'cad': {},  # Placeholder
    'document': {
        'docx': ['pdf', 'txt'],
        'pdf': ['jpg', 'png', 'txt'],
        'txt': ['pdf', 'docx']
    },
    'ebook': {
        'epub': ['pdf'],
        'mobi': ['pdf']
    },  # Limited support
    'font': {},  # Placeholder
    'image': {
        'png': ['jpg', 'ico', 'heic'],
        'jpg': ['png', 'ico', 'heic'],
        'ico': ['png', 'jpg'],
        'heic': ['png', 'jpg']
    },
    'other': {},  # Placeholder
    'presentation': {
        'pptx': ['pdf']
    },
    'spreadsheet': {
        'xlsx': ['csv', 'pdf']
    },
    'video': {
        'mp4': ['avi', 'mkv'],
        'avi': ['mp4', 'mkv'],
        'mkv': ['mp4', 'avi']
    }
}


def detect_file_type(file):
    mime = magic.Magic(mime=True)
    file.seek(0)
    file_data = file.read(1024)
    file.seek(0)
    mime_type = mime.from_buffer(file_data)
    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''

    if 'zip' in ext or 'application/zip' in mime_type:
        return 'archive', 'zip'
    elif 'rar' in ext or 'application/x-rar-compressed' in mime_type:
        return 'archive', 'rar'
    elif '7z' in ext or 'application/x-7z-compressed' in mime_type:
        return 'archive', '7z'
    elif 'audio/mpeg' in mime_type or ext == 'mp3':
        return 'audio', 'mp3'
    elif 'audio/wav' in mime_type or ext == 'wav':
        return 'audio', 'wav'
    elif 'audio/ogg' in mime_type or ext == 'ogg':
        return 'audio', 'ogg'
    elif 'pdf' in ext or 'application/pdf' in mime_type:
        return 'document', 'pdf'
    elif 'docx' in ext or 'application/vnd.openxmlformats-officedocument.wordprocessingml.document' in mime_type:
        return 'document', 'docx'
    elif 'txt' in ext or 'text/plain' in mime_type:
        return 'document', 'txt'
    elif 'epub' in ext:
        return 'ebook', 'epub'
    elif 'mobi' in ext:
        return 'ebook', 'mobi'
    elif 'png' in ext or 'image/png' in mime_type:
        return 'image', 'png'
    elif 'jpg' in ext or 'image/jpeg' in mime_type:
        return 'image', 'jpg'
    elif 'ico' in ext or 'image/x-icon' in mime_type:
        return 'image', 'ico'
    elif 'heic' in ext or 'image/heic' in mime_type:
        return 'image', 'heic'
    elif 'pptx' in ext or 'application/vnd.openxmlformats-officedocument.presentationml.presentation' in mime_type:
        return 'presentation', 'pptx'
    elif 'xlsx' in ext or 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in mime_type:
        return 'spreadsheet', 'xlsx'
    elif 'mp4' in ext or 'video/mp4' in mime_type:
        return 'video', 'mp4'
    elif 'avi' in ext or 'video/x-msvideo' in mime_type:
        return 'video', 'avi'
    elif 'mkv' in ext or 'video/x-matroska' in mime_type:
        return 'video', 'mkv'
    return 'other', ext


def convert_file(file, input_category, input_ext, output_ext):
    output = io.BytesIO()
    try:
        if input_category == 'archive':
            if input_ext == 'zip' and output_ext == 'rar':
                pass  # Placeholder
            elif input_ext == 'rar' and output_ext == 'zip':
                pass  # Placeholder
        elif input_category == 'audio':
            audio = AudioSegment.from_file(file, format=input_ext)
            audio.export(output, format=output_ext)
        elif input_category == 'document':
            if input_ext == 'pdf' and output_ext in ['jpg', 'png']:
                doc = fitz.open(stream=file.read(), filetype="pdf")
                page = doc.load_page(0)
                pix = page.get_pixmap()
                pix.save(output, output_ext)
            elif input_ext == 'docx' and output_ext == 'pdf':
                doc = Document(file)
                text = "\n".join([para.text for para in doc.paragraphs])
                pdf = fitz.open()
                page = pdf.new_page()
                page.insert_text((50, 50), text)
                pdf.save(output)
            elif input_ext == 'txt' and output_ext == 'pdf':
                text = file.read().decode('utf-8')
                pdf = fitz.open()
                page = pdf.new_page()
                page.insert_text((50, 50), text)
                pdf.save(output)
        elif input_category == 'image':
            img = Image.open(file)
            if output_ext == 'ico':
                img.save(output, format='ICO', sizes=[(16, 16), (32, 32)])
            else:
                img.convert('RGB').save(output, format=output_ext.upper())
        elif input_category == 'spreadsheet':
            if input_ext == 'xlsx' and output_ext == 'csv':
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                output.write(b','.join([str(cell.value or '') for cell in ws[1]]))
                for row in ws.iter_rows(min_row=2):
                    output.write(b'\n' + b','.join([str(cell.value or '') for cell in row]))
        elif input_category == 'video':
            video = mp.VideoFileClip(file)
            video.write_videofile(output, codec='libx264' if output_ext == 'mp4' else 'mpeg4')

        output.seek(0)
        return output, file.filename.rsplit('.', 1)[0] + '.' + output_ext
    except Exception as e:
        print(f"Error: {e}")
        return None, None